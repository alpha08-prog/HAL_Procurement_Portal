"""Asserts the approval layer against the documents it came from.

Same discipline as cascade_check.py: re-read the sources and compare, so the encoding
cannot drift quietly when the client reissues a file.

  * org.py       vs sampleData/Dummy HAL Database of Personnals.xlsx
  * checklist.py vs sampleData/Checklist for Indentor - STANDARD TERMS AND CONDITIONS-f.xlsx
                    -- including that every answer->authority mapping still has the
                       sheet's own words behind it
  * approval.py  vs the 14 hops the approved Provisioning Note prints on itself
                    -- the model has to represent the one real chain in sampleData

Run:  conda run -n hal python ai/approval_check.py
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import openpyxl

import approval
import checklist
import org
from tools.pdf_extractor import extract_routing_chain, scrub_watermark

F1_PDF = (Path(__file__).resolve().parents[1] / "sampleData" / "NOTING" /
          "NOTING SEQUENCE" / "F1 Approved Provisioning Note 6005612025.pdf")

# What the F1 note actually records, read off the document by hand once so the checks
# below have something independent to compare against.
F1_GRADES = [4, 6, 7, 7, 6, 7, 8, 8, 8, 7, 4, 7, 8, 9]
F1_HOPS = 14
F1_TWO_FACTOR = ["N4", "N8"]
F1_FIRST = ("SANJIB KUMAR RATH", "MANAGER (SECURITY)", "SECURITY")
F1_LAST = ("SUBRATA MONDAL", "GENERAL MANAGER (AOD)", "OFFICE OF GM (AOD)")

results = []


def check(ok, label, detail=""):
    results.append((bool(ok), label, detail))
    print(f"  {'ok  ' if ok else 'FAIL'} {label}" + (f"   -- {detail}" if detail and not ok else ""))
    return bool(ok)


def section(title):
    print(f"\n {title}")


# -- org.py vs the personnel sheet -------------------------------------------
def check_org():
    section("org.py -- the personnel directory")
    wb = openpyxl.load_workbook(org.XLSX, data_only=True, read_only=True)
    ws = wb[org.SHEET]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0] not in (None, "")]
    wb.close()

    ppl = org.load()
    check(len(ppl) == len(rows), f"every data row loaded ({len(ppl)} people)",
          f"sheet has {len(rows)}")
    check(len(ppl) == 1354, "1354 people", f"got {len(ppl)}")
    check(len({p.pb for p in ppl}) == len(ppl), "PB numbers are unique")
    check(min(p.pb for p in ppl) == "12345" and max(p.pb for p in ppl) == "13698",
          "PB range 12345-13698")
    check(len(org.divisions()) == 19, "19 units in the Division column",
          f"got {len(org.divisions())}")

    labels = {p.grade_label for p in ppl}
    unparsed = [l for l in labels if org.grade_level(l) is None]
    check(not unparsed, f"all {len(labels)} grade labels parse to a level",
          f"unparsed: {unparsed}")
    check(len(labels) == 23, "23 distinct grade labels", f"got {len(labels)}")

    # The sheet's own inconsistencies must not defeat the parser.
    check(org.grade_level("9 -  General Manager") == 9,
          "double space in \"9 -  General Manager\" still reads as 9")
    check(org.grade_level("1 - Assistant Finance Office") == 1,
          "the missing r in \"Assistant Finance Office\" is harmless")
    check(all(org.grade_level(k) == v for k, v in org.BOARD_TIER.items()),
          "CEO / SCH A / SCH B map above Executive Director")

    raw = {p.dept_raw for p in ppl}
    canon = {p.dept for p in ppl}
    check(len(raw) == 55, "55 raw department strings", f"got {len(raw)}")
    check(len(canon) == 47, "47 after normalising", f"got {len(canon)}")
    for a, b in [("Mat PLg", "Mat Plg"), ("SHOP", "Shop"), ("FIN", "Finance"),
                 ("MANUFACTURING SHHOP", "Manufacturing Shop"),
                 ("MARKETING", "Marketing"), ("TRANSPORT", "Transport"),
                 ("Project PLg", "PROJECT PLANNING")]:
        check(org.norm_dept(a) == org.norm_dept(b), f"\"{a}\" and \"{b}\" are one unit")

    # Designation -> grade, the bridge rules.LEVEL_DESIG needs to reach a person.
    for spec, lvl in [("GM(AOD)", 9), ("AGM(IMM-OH)", 8), ("DGM(FINANCE)", 7),
                      ("CM(Purchase)", 6), ("Manager (Security)", 4)]:
        check(org.desig_level(spec) == lvl, f"{spec} reads as grade {lvl}",
              f"got {org.desig_level(spec)}")
    check(org.desig_level("DY. GENERAL MANAGER (HR)") == 7,
          "\"DY. GENERAL MANAGER\" is 7, not 9 -- GM is a substring of DGM")
    check(org.desig_level("ADDL GENERAL MANAGER(PROJ & PLG)") == 8,
          "\"ADDL GENERAL MANAGER\" is 8, not 9")

    # Ambiguity is surfaced, never resolved silently.
    amb = sum(1 for d, depts in org.unit_tree().items()
              for x in depts if org.head_of(d, x)["ambiguous"])
    total = sum(len(v) for v in org.unit_tree().values())
    check(amb > 0, f"{amb} of {total} division-department pairs have no single top grade",
          "expected the sheet to be ambiguous -- it has no head column")
    h = org.head_of("DIV1", "IMM")
    check(h["ambiguous"] and h["person"] is None and len(h["candidates"]) == 3,
          "head_of returns candidates, not a guess, when the top grade is tied")
    picked, chose = org.pick(h)
    check(picked is not None and chose,
          "org.pick breaks a tie but reports that it did")

    # Unit-hint matching.
    check(org.match_dept("IMM-OH", "DIV1") == "IMM", "\"IMM-OH\" narrows to IMM")
    check(org.match_dept("SEC & FIRE", "DIV1") == "FIRE & SEC",
          "\"SEC & FIRE\" matches FIRE & SEC on its words")
    check(org.match_dept("AOD", "DIV1") == "",
          "\"AOD\" is a division, not a department -- no false match")


# -- checklist.py vs the checklist sheet --------------------------------------
def check_checklist():
    section("checklist.py -- the indentor checklist")
    rows = checklist.rows()
    check(len(rows) == 67, f"67 rows (sheet rows 4-70)", f"got {len(rows)}")
    check(len(checklist.block(checklist.PROVISIONING)) == 25, "25 provisioning rows")
    check(len(checklist.block(checklist.TENDER)) == 42, "42 tender rows")

    t, c = len(checklist.by_category("T")), len(checklist.by_category("C"))
    check(t == 42, "42 rows marked T (technical)", f"got {t}")
    check(c == 21, "21 rows marked C (commercial)", f"got {c}")
    check(len(rows) - t - c == 4, "4 rows carry no category")

    tec = len(checklist.consumed_by("tender+tec_report"))
    ce = len(checklist.consumed_by("tender+comm_eval"))
    ind = len(checklist.consumed_by("indentor"))
    check(tec == 15, "15 rows feed the TEC Report (column H)", f"got {tec}")
    check(ce == 27, "27 rows feed the Commercial Evaluation (column H)", f"got {ce}")
    check(ind == 25, "25 rows are indentor-only", f"got {ind}")
    check(tec + ce + ind == len(rows), "every row's column H is recognised",
          "some rows fell through to 'unknown'")

    # Column H is matched verbatim, typos included, so a reissue that rewords it fails.
    wb = openpyxl.load_workbook(checklist.XLSX, data_only=True)
    ws = wb[checklist.SHEET]
    sheet_h = {" ".join(str(ws[f"H{r}"].value).split())
               for r in range(4, 71) if ws[f"H{r}"].value}
    wb.close()
    for key, text in checklist.CONSUMED.items():
        check(" ".join(text.split()) in sheet_h,
              f"column H text for \"{key}\" still appears in the sheet verbatim")

    check(len(checklist.material_classes()) == 13, "13 material classes on Sheet2")
    check(checklist.dop_level() == "Level I",
          "DOP level reads as Level I from provisioning sl 11",
          f"got {checklist.dop_level()}")

    # The tender block's clause numbering shifted by one when the June-2026 revision
    # inserted "In Case of Three Bid Pre Meeting" as sl 5. Pin the rows that carry
    # money rules, because ARCHITECTURE.md quotes them by number.
    for sl, clause in [("5", "In Case of Three Bid Pre Meeting"),
                       ("17", "Equipment Warranty"), ("18", "Liquidated damages"),
                       ("21", "Validity of Offer"), ("28", "Security Deposit"),
                       ("29", "Performance Bank Guarantee")]:
        row = checklist.find(checklist.TENDER, sl)
        check(row is not None and row["clause"].upper().startswith(clause.upper()[:14]),
              f"tender sl {sl} is \"{clause}\"",
              f"got {row['clause'] if row else 'missing'}")


# -- the answer -> authority table -------------------------------------------
def check_injections():
    section("checklist.INJECTIONS -- who the answers pull into the chain")
    check(len(checklist.INJECTIONS) == 10, "10 rows name an approving authority",
          f"got {len(checklist.INJECTIONS)}")

    for spec in checklist.INJECTIONS:
        row = checklist.find(spec["block"], spec["sl"])
        if not check(row is not None, f"{spec['block']} sl {spec['sl']} exists in the sheet"):
            continue
        text = checklist.evidence_text(row).upper()
        check(spec["evidence"].upper() in text,
              f"sl {spec['sl']}: the sheet still says \"{spec['evidence']}\"",
              f"not found in: {text[:120]}")

    neg = [s for s in checklist.INJECTIONS if s["trigger"] == "negative"]
    check(len(neg) == 1 and neg[0]["sl"] == "13",
          "sl 13 is the only inverted trigger "
          "(\"not raised within six months\" -- a YES means compliant)")

    ans = checklist.answers()
    fired = {i["sl"] for i in checklist.injected(ans)}
    check(fired == {"10", "11", "12", "21"},
          "on the sheet's own answers, four authorities are pulled in (sl 10, 11, 12, 21)",
          f"got {sorted(fired)}")
    check("13" not in fired,
          "sl 13 answered YES pulls in nobody -- the inverted trigger works")
    check("19" not in fired, "sl 19 answered NA does not summon the Ministry")

    # Flip the answers and the chain must change.
    flipped = dict(ans, **{"provisioning:22": "YES", "provisioning:13": "NO"})
    f2 = {i["sl"] for i in checklist.injected(flipped)}
    check("22" in f2, "answering YES to short tendering adds the Head of Division")
    check("13" in f2, "answering NO to the six-month rule adds the Head of Division")

    ext = [i for i in checklist.injected(dict(ans, **{"provisioning:19": "YES"}))
           if i["kind"] == "ministry"]
    check(ext and ext[0]["external"],
          "the Ministry is flagged external -- no directory entry can satisfy it")


# -- the model vs the real chain ---------------------------------------------
def check_replay():
    section("approval.py -- against the 14 hops the F1 note prints on itself")
    if not check(F1_PDF.exists(), f"{F1_PDF.name} is present"):
        return

    hops = extract_routing_chain(str(F1_PDF))
    check(len(hops) == F1_HOPS, f"{F1_HOPS} hops extracted", f"got {len(hops)}")

    grades = [org.desig_level(h["designation"]) for h in hops]
    check(grades == F1_GRADES, f"grade path is {F1_GRADES}", f"got {grades}")
    check(not all(b >= a for a, b in zip(grades, grades[1:])),
          "the real chain is NOT grade-monotonic -- so hop order must never be "
          "validated by grade")

    first, last = hops[0], hops[-1]
    check((first["name"], first["designation"], first["dept"]) == F1_FIRST,
          "N1 is the originator, Manager (Security)",
          f"got {(first['name'], first['designation'], first['dept'])}")
    check((last["name"], last["designation"], last["dept"]) == F1_LAST,
          "N14 is GM(AOD), the CFA",
          f"got {(last['name'], last['designation'], last['dept'])}")

    tfa = [h["note"] for h in hops if h["two_factor"]]
    check(tfa == F1_TWO_FACTOR, f"two-factor on {F1_TWO_FACTOR} only", f"got {tfa}")

    # Watermark scrub.
    check(scrub_watermark("9 5 8 8 HIRALAL KESHI") == "HIRALAL KESHI",
          "scrub_watermark strips interleaved print-ID digits")
    check(scrub_watermark("2 2 1 18/07/2025") == "18/07/2025", "dates survive the scrub")
    check(scrub_watermark("8 3 6 N10") == "N10", "note ids survive the scrub")
    check(scrub_watermark("0 1 11") == "11", "multi-digit serials survive the scrub")
    check(scrub_watermark("7") == "7", "a legitimate single-digit cell is left alone")
    dirty = [h["note"] for h in hops
             if any(len(tok) == 1 and tok.isdigit()
                    for f in ("name", "designation", "dept", "division", "date")
                    for tok in h[f].split())]
    check(not dirty, "no watermark digits survive in any hop field", f"dirty: {dirty}")

    # The remark HAL fills in by default, verbatim including the Hindi half.
    concurs = [h for h in hops if approval.CONCUR_DEFAULT in h["comment"]]
    check(len(concurs) == 4,
          f"approval.CONCUR_DEFAULT appears verbatim on 4 hops (N3, N4, N6, N7)",
          f"got {len(concurs)}")
    check(approval.clean_comment("") == approval.CONCUR_DEFAULT,
          "an empty remark becomes the standard line")
    check(approval.clean_comment(" . * , ") == approval.CONCUR_DEFAULT,
          "a symbols-only remark becomes the standard line")
    check(approval.clean_comment("Pl examine") == "Pl examine",
          "a real remark is left alone")

    # Every hop type the real note needs must exist in the vocabulary.
    for h in ("forward", "concur", "concur_with_rider", "examine", "query",
              "approve", "reject", "return_to"):
        check(h in approval.HOPS, f"hop type \"{h}\" is defined")
    check(approval.HOPS["examine"]["advances"] is False,
          "\"examine\" does not advance the chain -- it comes back (F1 N9->N10)")
    check(approval.HOPS["query"]["advances"] is False,
          "\"query\" does not advance the chain -- N10's objection kept its place")
    check(approval.HOPS["approve"]["by"] == "cfa", "only the CFA may approve")

    # Rebuild the chain and prove the gate is the thing that holds the file.
    plan = approval.build_plan("provisioning", "DIV9", originator_dept="FIRE & SEC")
    chain = approval.Chain(plan, file_id="6005612025")
    for h in hops:
        person = org.Person(f"F1-{h['seq']:02d}", h["name"], h["division"],
                            h["dept"], h["designation"], h["designation"])
        person.grade_level = org.desig_level(h["designation"])
        act = "approve" if h["seq"] == len(hops) else "concur"
        slot = next((s for s in plan.slots
                     if s.kind != "originator" and not s.actioned), None)
        chain.add(person, act, h["comment"], slot=slot, when=h["date"],
                  two_factor=h["two_factor"])
    check(len(chain.hops) == F1_HOPS, "the model holds all 14 hops")
    check(chain.grade_path() == F1_GRADES,
          "the rebuilt chain keeps the non-monotonic grade path")
    check(chain.decision == "approve", "the rebuilt chain ends approved")
    # 18/07/2025 -> 21/08/2025: 13 days left in July plus 21 in August.
    check(chain.elapsed_days() == 34, "34 days from N1 to N14",
          f"got {chain.elapsed_days()}")
    check(len({h["txn_id"] for h in chain.hops}) == F1_HOPS,
          "every hop gets its own transaction id, as the real note does")
    check(chain.hops[0]["txn_id"].endswith("-0001")
          and chain.hops[-1]["txn_id"].endswith("-000E"),
          "the id counter runs 0001..000E, matching the note")

    # The gate: approval alone is not enough while a required slot is idle.
    plan2 = approval.build_plan("provisioning", "DIV9", originator_dept="FIRE & SEC")
    bare = approval.Chain(plan2, file_id="X")
    if plan2.cfa and plan2.cfa.person:
        bare.add(plan2.cfa.person, "approve", "Approved", slot=plan2.cfa)
    ok, why = bare.release_ready()
    check(not ok and why,
          "a CFA approval with concurrences outstanding does NOT release the file",
          "the gate let it through")

    full = approval.Chain(plan2, file_id="Y")
    for s in plan2.slots:
        if s.person is None or s.kind in ("originator", "cfa"):
            continue
        full.add(s.person, "concur", "", slot=s)
    full.add(plan2.cfa.person, "approve", "Approved / मंजूर", slot=plan2.cfa)
    ok2, why2 = full.release_ready()
    check(ok2, "with every required slot actioned and the CFA approved, it releases",
          f"still blocked by: {why2}")


# -- committee mode ----------------------------------------------------------
def check_committee():
    section("approval.Committee -- stage 3 is a committee, not a chain")
    check("no conflict of interest" in approval.COI_DECLARATION,
          "the Annexure 21A Para C declaration is carried verbatim")
    com = approval.Committee("tec_report", "test")
    ok, why = com.complete()
    check(not ok and "no members named" in why[0], "an empty committee cannot report")

    p = org.load()[0]
    com.add_member(p, spec="AGM(QA) - Chairman")
    ok, why = com.complete()
    check(not ok, "an unsigned member blocks the report")
    com.sign(0, coi_declared=False)
    ok, why = com.complete()
    check(not ok and "conflict-of-interest" in why[0],
          "a member who signed but did not declare still blocks the report")
    com.sign(0, coi_declared=True)
    ok, _ = com.complete()
    check(ok, "signed and declared -- the report can be raised")

    shape = approval.chain_shape("tec_report")
    check(shape["mode"] == approval.COMMITTEE, "tec_report is committee mode")
    check(not shape.get("committee_specs"),
          "TEC composition is left empty -- it is not in sampleData, so it is not invented")
    check(approval.chain_shape("pnc_req").get("committee_from_case") == ("pnc", "committee"),
          "the PNC committee is seeded from F5 via case_input, where it IS named")


def main():
    print("=" * 70)
    print(" approval layer -- assertions against the source documents")
    print("=" * 70)
    check_org()
    check_checklist()
    check_injections()
    check_replay()
    check_committee()

    passed = sum(1 for ok, _, _ in results if ok)
    total = len(results)
    print("\n" + "=" * 70)
    print(f" {passed}/{total} checks passed")
    print("=" * 70)
    if passed != total:
        for ok, label, detail in results:
            if not ok:
                print(f"  FAIL {label}" + (f"   -- {detail}" if detail else ""))
    return 0 if passed == total else 1


if __name__ == "__main__":
    sys.exit(main())
