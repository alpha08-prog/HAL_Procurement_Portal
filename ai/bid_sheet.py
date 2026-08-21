"""Read filled technical-bid compliance sheets and derive the case facts from them.

    conda run -n hal python ai/bid_sheet.py

HAL's technical-bid compliance sheet is the return leg of a tender: the bidder writes
YES or NO against every specification line and every term, states its Udyam number, NIC
code and EMD position, and fills in the vendor block. Two of the pipeline's stages read
nothing else --

  * the **EMD Stage Acceptance Note** turns on one question per bidder: is the waiver
    claim valid, i.e. does this bidder *manufacture* the offered product in the relevant
    NIC category? (rules.emd_waiver)
  * the **TEC Report** rejects offers by citing the specification sl nos they failed,
    which is exactly the NO rows of this sheet.

So this module is the bridge from the sheet to `case_input.json`. It reads any workbook
laid out like the client's form -- one sheet per bidder plus a PRICE BIDS sheet -- and
returns the `bidders`, `tec` and `price_bid` sections the pipeline expects, with the
numbers computed in code rather than transcribed.

Provenance is kept explicit: the fixture it ships against is fabricated (see
ai/fixtures/make_bid_E33046.py), and the case file it writes says so in `_provenance`.
`ai/case_input.json` is untouched -- that one is seeded from sampleData only.
"""

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import openpyxl

import rules

HERE = Path(__file__).parent
DEFAULT_BOOK = HERE / "fixtures" / "TechnicalBid_E-33046_FILLED.xlsx"
DEFAULT_OUT = HERE / "fixtures" / "case_input_E33046.json"

SKIP_SHEETS = {"READ ME", "PRICE BIDS"}

# NIC codes that count as manufacturing the offered product. 27400 is "manufacture of
# electric lighting equipment"; 465xx is wholesale trade, which is what makes a
# trader's MSE waiver claim fail -- the same test the NVB case applied.
MANUFACTURING_NIC = ("27400", "2740", "274")


def _s(v):
    return "" if v is None else " ".join(str(v).split())


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = re.sub(r"[^\d.]", "", str(v))
    return float(t) if t else None


def _find_label(ws, needle, value_col=3, limit=400):
    """Value beside the first row whose column B contains `needle`."""
    n = needle.lower()
    for r in range(1, min(ws.max_row, limit) + 1):
        if n in _s(ws.cell(row=r, column=2).value).lower():
            return _s(ws.cell(row=r, column=value_col).value)
    return ""


def read_bidder(ws):
    """One bidder's returned sheet."""
    name = ""
    for r in range(1, 8):
        t = _s(ws.cell(row=r, column=1).value)
        if t.upper().startswith("NAME OF THE BIDDER"):
            name = t.split(":", 1)[1].strip() if ":" in t else ""
    bid = {
        "id": ws.title,
        "name": name or ws.title,
        "udyam": _find_label(ws, "Udyam Registration"),
        "nic": _find_label(ws, "NIC code"),
        "emd_claim": _find_label(ws, "EMD status"),
        "nature": _find_label(ws, "Nature of Firm"),
        "msme": _find_label(ws, "MSME"),
        "mfr_class": _find_label(ws, "Classification of Manufacturer"),
        "gst": _find_label(ws, "GST No"),
        "pan": _find_label(ws, "PAN No"),
        "email": _find_label(ws, "e-Mail Address"),
        "address": _find_label(ws, "Registered Address"),
        "spec_failed": [],
        "spec_remarks": [],
        "terms_failed": [],
    }

    # Walk the specification block: sl no in A, compliance in E, remark in F.
    in_specs = False
    for r in range(1, ws.max_row + 1):
        a = _s(ws.cell(row=r, column=1).value)
        if a.upper().startswith("TECHNICAL SPECIFICATIONS"):
            in_specs = True
            continue
        if a.upper().startswith("OTHER APPLICABLE TERMS"):
            in_specs = False
            continue
        if not in_specs or not a.isdigit():
            continue
        if _s(ws.cell(row=r, column=5).value).upper().startswith("NO"):
            bid["spec_failed"].append(int(a))
            bid["spec_remarks"].append(
                f"sl no {a}: {_s(ws.cell(row=r, column=6).value)}")

    # Terms block: compliance moves to column D, remark to E.
    in_terms = False
    for r in range(1, ws.max_row + 1):
        a = _s(ws.cell(row=r, column=1).value)
        if a.upper().startswith("OTHER APPLICABLE TERMS"):
            in_terms = True
            continue
        if a.upper().startswith("VENDOR DETAILS"):
            in_terms = False
            continue
        if not in_terms or not a.isdigit():
            continue
        if _s(ws.cell(row=r, column=4).value).upper().startswith("NO"):
            bid["terms_failed"].append(int(a))

    # The EMD decision, computed -- never read off the sheet. A bidder who paid is in;
    # a bidder claiming a waiver is in only if it manufactures in the relevant category.
    claim = bid["emd_claim"].lower()
    paid = "paid" in claim and "not paid" not in claim
    seeks_waiver = "waiver" in claim
    is_mfr = any(k in bid["nature"].lower() for k in ("manufacturer", "oem"))
    nic_ok = bid["nic"].startswith(MANUFACTURING_NIC)
    bid["manufacturer"], bid["nic_match"] = is_mfr, nic_ok
    if paid:
        bid["emd"], bid["emd_reason"] = "Accepted", "EMD paid through SB Collect"
    elif seeks_waiver and rules.emd_waiver(bid):
        bid["emd"] = "Accepted"
        bid["emd_reason"] = (f"Waiver valid - manufacturer of the offered product in "
                             f"NIC {bid['nic']}")
    elif seeks_waiver:
        bid["emd"] = "Not Accepted"
        why = ("not a manufacturer of the offered product"
               if not is_mfr else f"NIC {bid['nic']} is not the relevant category")
        bid["emd_reason"] = f"Waiver claim rejected - {why}"
    else:
        bid["emd"] = "Not Accepted"
        bid["emd_reason"] = "EMD neither paid nor validly waived"
    return bid


def read_prices(ws):
    """The price-bid envelope: per-bidder rates plus estimate, LPP and counter-offer."""
    prices, extras = {}, {}
    header = None
    for r in range(1, ws.max_row + 1):
        a = _s(ws.cell(row=r, column=1).value)
        if a == "Bidder":
            header = r
            continue
        if header is None or not a:
            continue
        unit = _num(ws.cell(row=r, column=4).value)
        landed = _num(ws.cell(row=r, column=7).value)
        if unit is None:
            if a.lower().startswith("reverse auction"):
                extras["ra_status"] = _s(ws.cell(row=r, column=2).value)
            elif a.lower().startswith("lpp contract"):
                extras["lpp_contract"] = _s(ws.cell(row=r, column=2).value)
            continue
        row = {"unit_basic": unit, "basic": _num(ws.cell(row=r, column=5).value),
               "landed": landed}
        low = a.lower()
        if low.startswith("provisioning estimate"):
            extras["estimate"] = row
        elif low.startswith("last purchase"):
            extras["lpp"] = row
        elif low.startswith("counter-offer"):
            extras["counter"] = row
        else:
            prices[a] = row
    return prices, extras


def load(book=None):
    p = Path(book) if book else DEFAULT_BOOK
    if not p.exists():
        raise FileNotFoundError(
            f"{p} not found -- generate it first:\n"
            f"  conda run -n hal python ai/fixtures/make_bid_E33046.py")
    wb = openpyxl.load_workbook(p, data_only=True)
    bidders = [read_bidder(wb[s]) for s in wb.sheetnames if s not in SKIP_SHEETS]
    prices, extras = read_prices(wb["PRICE BIDS"]) if "PRICE BIDS" in wb.sheetnames else ({}, {})
    tender_ref = ""
    for s in wb.sheetnames:
        if s in SKIP_SHEETS:
            continue
        for r in range(1, 6):
            t = _s(wb[s].cell(row=r, column=1).value)
            if t.upper().startswith("TENDER REF"):
                tender_ref = t.split(":", 1)[-1].strip()
                break
        if tender_ref:
            break
    wb.close()
    for b in bidders:
        b["price"] = prices.get(b["id"], {})
    return {"tender_ref": tender_ref, "bidders": bidders, "extras": extras}


def to_case_input(data, meta):
    """Assemble the pipeline's input sections from the sheets.

    Every figure here is computed by rules.py from what the bidders wrote -- the
    variance, the savings, SD and PBG. Nothing is transcribed.
    """
    bidders, ex = data["bidders"], data["extras"]
    emd_ok = [b for b in bidders if b["emd"] == "Accepted"]
    # The TEC sees only the EMD-accepted offers, and rejects those with a NO row.
    tec_rejected = [b for b in emd_ok if b["spec_failed"]]
    tec_accepted = [b for b in emd_ok if not b["spec_failed"]]

    priced = [b for b in tec_accepted if b["price"].get("landed")]
    priced.sort(key=lambda b: b["price"]["landed"])
    l1 = priced[0] if priced else None

    est = (ex.get("estimate") or {}).get("landed")
    l1_landed = l1["price"]["landed"] if l1 else None
    counter = (ex.get("counter") or {}).get("landed")
    lpp = (ex.get("lpp") or {}).get("landed")

    variance = rules.variance(l1_landed, est)
    sav_amt, sav_pct = rules.savings(l1_landed, counter)
    basic = rules.basic_of(counter) if counter else None

    return {
        "_provenance": (
            "DERIVED by ai/bid_sheet.py from ai/fixtures/TechnicalBid_E-33046_FILLED.xlsx. "
            "THE BIDS ARE FABRICATED TEST DATA -- see ai/fixtures/make_bid_E33046.py. "
            "The tender reference, item, quantity, 12 specification lines and 18 terms "
            "come verbatim from sampleData/TechnicalBid E-33046.pdf; the bidders, their "
            "compliance answers and every price are invented. This file is a FIXTURE and "
            "is NOT ai/case_input.json, which is seeded from sampleData only."),
        "_fixture": True,
        "_tender_ref": data["tender_ref"],
        "requisition": {
            "item_description": meta["item"], "car_no": meta["car_no"],
            "car_date": meta["car_date"], "reference_no": meta["reference_no"],
            "quantity": meta["qty"], "budget_year": meta["budget_year"],
            "budget_type": meta["budget_type"],
            "mpr_estimate": f"{est:,.0f}" if est else None,
            "amount_in_words": meta["amount_in_words"],
            "dop_clause": meta["dop_clause"],
        },
        "tender": {
            "tender_no": meta["tender_no"], "tender_date": meta["tender_date"],
            "tender_enquiry": meta["enquiry_no"], "tender_type": meta["tender_type"],
            "total_bids": len(bidders),
        },
        "bidders": [{"name": b["name"], "udyam": b["udyam"], "mse": b["msme"],
                     "nic": b["nic"], "emd": b["emd"], "emd_reason": b["emd_reason"],
                     "nature": b["nature"]} for b in bidders],
        "tec": {
            "query": ("TEC sought confirmation of BIS CRS registration and photometric "
                      "test reports from the bidders; replies received and re-evaluated."),
            "pm_clause": "8.5.6",
            "accepted": [b["name"] for b in tec_accepted],
            "rejected": [{"name": b["name"],
                          "spec_slnos": ", ".join(str(s) for s in b["spec_failed"])}
                         for b in tec_rejected],
        },
        "price_bid": {
            "l1_vendor": l1["name"] if l1 else None,
            "l1_price": f"{l1_landed:,.0f}" if l1_landed else None,
            "ra_status": ex.get("ra_status"),
            "cst_ref": "Annexure 21B",
        },
        "lpp": {"contract": ex.get("lpp_contract"), "date": "22.08.2024",
                "price": f"{lpp:,.0f}" if lpp else None, "model": "LED-HB-250"},
        "counter_offer": {"price": f"{counter:,.0f}" if counter else None,
                          "date": meta["counter_date"],
                          "savings_amount": f"{sav_amt:,.0f}" if sav_amt else None},
        "pnc": {"committee": meta["pnc_committee"]},
        "proposal": {
            "id": meta["proposal_id"], "vendor": l1["name"] if l1 else None,
            "initiator": meta["initiator"], "initiator_desig": meta["initiator_desig"],
            "fca": meta["fca"], "fca_desig": meta["fca_desig"],
            "cfa": meta["cfa"], "cfa_desig": meta["cfa_desig"],
            "dop_level": meta["dop_level"],
            "value": f"{counter:,.0f}" if counter else None,
        },
        "_computed": {
            "price_variance_pct": variance, "savings_pct": sav_pct,
            "sd_amount": rules.sd(basic), "pbg_amount": rules.pbg(basic),
            "note": "computed by rules.py from the sheets, not transcribed",
        },
    }


# Case facts that are not on a bid sheet -- the indent side. Fabricated, like the bids.
META = {
    "item": "250W HIGH BAY LED LIGHT FITTING",
    "car_no": "CAR/26/118", "car_date": "11 May 2026",
    "reference_no": "PM/1989/LED/2026/044", "qty": 2150,
    "budget_year": "2026-27", "budget_type": "Revenue Budget",
    "amount_in_words": "One Crore Six Lakh Fifty Five Thousand Four Hundred Only",
    "dop_clause": "Annexure III B, Sl No 1a",
    "tender_no": "GEM/2026/B/7729104", "tender_date": "05-06-2026",
    "enquiry_no": "E-33046", "tender_type": "GeM Custom Bid (Open), Two Bid, Line-wise",
    "counter_date": "14.07.2026",
    "pnc_committee": ["AGM(Fin) - Chairman (senior-most & Finance)",
                      "AGM(IMM) - Member Secretary",
                      "DGM(Plant Maint.) - Member (user)",
                      "DGM(Purchase) - Member"],
    "proposal_id": "PP/AOD/IMM/2026/0331",
    "initiator": "<from the personnel directory>", "initiator_desig": "Manager (IMM)",
    "fca": "<Head of Finance>", "fca_desig": "AGM (Finance)",
    "cfa": "<Divisional Head Level-I>", "cfa_desig": "GM (AOD)",
    "dop_level": "Level I  [human-supplied: DOP-2025 value bands are not in sampleData]",
}


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Read filled technical-bid compliance sheets into pipeline input")
    ap.add_argument("--book", default=str(DEFAULT_BOOK), help="the filled workbook")
    ap.add_argument("--out", default=str(DEFAULT_OUT), help="case-input JSON to write")
    ap.add_argument("--quiet", action="store_true")
    a = ap.parse_args(argv)

    data = load(a.book)
    ci = to_case_input(data, META)

    if not a.quiet:
        print("=" * 76)
        print(f" TECHNICAL BID COMPLIANCE SHEETS -- {data['tender_ref']}")
        print("=" * 76)
        print(f" {len(data['bidders'])} bids read from {Path(a.book).name}\n")
        print(f" {'bid':<5}{'name':<34}{'nature':<20}{'NIC':<8}{'EMD':<14}spec NOs")
        for b in data["bidders"]:
            fails = ", ".join(str(s) for s in b["spec_failed"]) or "-"
            print(f" {b['id']:<5}{b['name'][:32]:<34}{b['nature'][:18]:<20}"
                  f"{b['nic']:<8}{b['emd']:<14}{fails}")

        print(f"\n{'-' * 76}\n EMD STAGE -- computed by rules.emd_waiver, not read off the sheet\n{'-' * 76}")
        for b in data["bidders"]:
            print(f"  {b['id']}  {b['emd']:<14}{b['emd_reason']}")

        print(f"\n{'-' * 76}\n TEC STAGE -- rejections cite the sl nos the bidder answered NO to\n{'-' * 76}")
        print(f"  accepted: {', '.join(ci['tec']['accepted']) or 'none'}")
        for r in ci["tec"]["rejected"]:
            who = next(b for b in data["bidders"] if b["name"] == r["name"])
            print(f"  rejected: {r['name']}  -- not complied, sl no {r['spec_slnos']}")
            for rm in who["spec_remarks"]:
                print(f"              {rm}")

        print(f"\n{'-' * 76}\n PRICE BID & NEGOTIATION -- every figure from rules.py\n{'-' * 76}")
        c = ci["_computed"]
        print(f"  estimate (landed)  : {ci['requisition']['mpr_estimate']}")
        print(f"  L1                 : {ci['price_bid']['l1_vendor']} at "
              f"{ci['price_bid']['l1_price']}")
        print(f"  variance vs estimate: {c['price_variance_pct']}%")
        print(f"  RA                 : {ci['price_bid']['ra_status']}")
        print(f"  pnc_required()     : "
              f"{rules.pnc_required({'l1_price': ci['price_bid']['l1_price'], 'budget_estimate': ci['requisition']['mpr_estimate'], 'ra_status': ci['price_bid']['ra_status']})}")
        print(f"  counter-offer      : {ci['counter_offer']['price']}  "
              f"(saving {ci['counter_offer']['savings_amount']} = {c['savings_pct']}%)")
        print(f"  SD 5% / PBG 10%    : {c['sd_amount']:,.2f} / {c['pbg_amount']:,.2f}")

    Path(a.out).write_text(json.dumps(ci, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n case input -> {Path(a.out).relative_to(HERE.parent)}")
    print(" FIXTURE: the bids are fabricated. ai/case_input.json is untouched.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
