"""The Indentor Checklist as data -- a machine-readable copy of

    sampleData/Checklist for Indentor - STANDARD TERMS AND CONDITIONS-f.xlsx
        [Sheet1 rows 4-70, Sheet2]

Two blocks, 67 rows:

  rows  4-28   the PROVISIONING block -- sl A/B/C then 1..22. Becomes the
               Provisioning Note. Column H: "filled by Indnetor".
  rows 29-70   the TENDER block -- sl 1..42. Becomes the Tender Document, and
               column H says which downstream document each row also feeds:
               15 rows -> the TEC Report, 27 rows -> the Commercial Evaluation.

Column G (`category (tech/Commercial)`) marks each row T or C: 42 technical,
21 commercial, 4 blank.

The part that matters for the approval chain: **nine rows name an approving
authority in their own text.** Answering "yes" to short tendering pulls in the Head
of Division; answering "yes" to brand-specific procurement pulls in a Committee;
the global-tender-exemption row reaches outside HAL entirely, to the Ministry. So
who signs a Provisioning Note is decided by how this sheet was filled in, not by a
fixed ladder. INJECTIONS below is that mapping, each entry carrying the sheet's own
words as evidence so approval_check.py can prove it has not drifted.

The sheet's typos ("Indnetor", "Evalaution", "obatined", "Provisioining") are matched
verbatim rather than corrected, the same discipline cascade_check.py uses.
"""

import re
import sys
from functools import lru_cache
from pathlib import Path

import openpyxl

XLSX = (Path(__file__).resolve().parents[1] / "sampleData" /
        "Checklist for Indentor - STANDARD TERMS AND CONDITIONS-f.xlsx")

SHEET, CLASS_SHEET = "Sheet1", "Sheet2"

FIRST_ROW, BLOCK_SPLIT, LAST_ROW = 4, 29, 70

PROVISIONING, TENDER = "provisioning", "tender"

# Column H, verbatim. Anything else is reported as unknown rather than guessed at.
CONSUMED = {
    "indentor": "Check list Part to be filled by Indnetor",
    "tender+comm_eval": ("Check list Part to be filled by Indnetor & used in Tender doc "
                         "creation & commercial Evalaution by tendering Group"),
    "tender+tec_report": ("Check list Part to be filled by Indnetor & used in Tender doc "
                          "creation & TEC report"),
}

# -- the answer -> authority table -------------------------------------------
# kind      what approval.py has to resolve
# trigger   "affirmative" injects when the answer is a yes; "negative" injects when it
#           is NOT a yes -- sl 13 reads "Same requirement *not* Raised within Six
#           Months", so a "YES" there means compliant and needs nobody extra, while
#           anything else is a splitting risk that needs the Head of Division.
# evidence  a substring that must still appear in the row's own clause/description/
#           remark text, so this table cannot silently drift from the sheet.
INJECTIONS = [
    {"block": PROVISIONING, "sl": "10", "kind": "cfa",
     "authority": "Provisioning CFA per DOP-2025 & PM Iss-4", "trigger": "affirmative",
     "evidence": "CFA Name, Grade & Designation"},
    {"block": PROVISIONING, "sl": "11", "kind": "dop_authority",
     "authority": "Provisioning Authority under DOP-2025", "trigger": "affirmative",
     "evidence": "Approval from Provisioning Authority obtained under DOP-2025"},
    {"block": PROVISIONING, "sl": "12", "kind": "dop_authority",
     "authority": "DOP-2025 authority for proprietary / single tender",
     "trigger": "affirmative", "evidence": "as per DOP-2025"},
    {"block": PROVISIONING, "sl": "13", "kind": "head_of_division",
     "authority": "Head of Division (requirement repeated within six months)",
     "trigger": "negative", "evidence": "Permission from Head of Division"},
    {"block": PROVISIONING, "sl": "15", "kind": "head_of_division",
     "authority": "Head of Division (e-tendering waiver above INR 2.00 lakhs)",
     "trigger": "affirmative", "evidence": "waiver from Head of Division"},
    {"block": PROVISIONING, "sl": "16", "kind": "head_of_division",
     "authority": "Head of Division (limited tender below 5 sources, PM para 6.8.2)",
     "trigger": "affirmative", "evidence": "exemption to be obtained from Head of Division"},
    {"block": PROVISIONING, "sl": "18", "kind": "committee",
     "authority": "Committee / CPA (brand or make specific procurement)",
     "trigger": "affirmative", "evidence": "Approval obtained from respective Committee/CPA"},
    {"block": PROVISIONING, "sl": "19", "kind": "ministry",
     "authority": "Concerned Ministry (global tender exemption)", "trigger": "affirmative",
     "evidence": "permission is required from concerned Ministry", "external": True},
    {"block": PROVISIONING, "sl": "21", "kind": "indigenisation_cell",
     "authority": "Indigenisation Cell (project item from a non-local supplier)",
     "trigger": "affirmative", "evidence": "confirmation from Indigenization cell"},
    {"block": PROVISIONING, "sl": "22", "kind": "head_of_division",
     "authority": "Head of Division (short tender, under three weeks TOD)",
     "trigger": "affirmative", "evidence": "from Head of Division for Short Tendering"},
]

# Answers that read as "not applicable / nothing sought here".
_NOT_APPLICABLE = {"", "NA", "N/A", "NIL", "NO", "NOT APPLICABLE", "-"}


def _s(v):
    return "" if v is None else " ".join(str(v).split())


def is_affirmative(answer):
    """Does this answer read as a yes?

    The sheet's compliance column is a free-ish dropdown: "YES", "Yes",
    "YES-Included in MPR", "Yes-CPA is Level I", "Clause to be included in Tender T&C",
    "NA", "Composite", "Goods". Only an explicit yes counts as one.
    """
    t = _s(answer).upper()
    if t in _NOT_APPLICABLE:
        return False
    return t.startswith("YES") or t == "Y"


def _consumed_of(h):
    t = _s(h)
    for key, text in CONSUMED.items():
        if t == _s(text):
            return key
    return "unknown" if t else ""


@lru_cache(maxsize=1)
def rows(path=None):
    """Rows 4-70 as dicts, in sheet order."""
    p = Path(path) if path else XLSX
    if not p.exists():
        raise FileNotFoundError(f"indentor checklist not found: {p}")
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb[SHEET]
    out = []
    for r in range(FIRST_ROW, LAST_ROW + 1):
        sl = _s(ws[f"A{r}"].value)
        if not sl:
            continue
        out.append({
            "row": r,
            # The block is the row range, not column B: row 43 ("Site Readiness") is
            # labelled "For Provisioning File" in B but carries tender sl 15 and feeds
            # the TEC report, so B cannot be trusted to split the blocks.
            "block": PROVISIONING if r < BLOCK_SPLIT else TENDER,
            "sl": sl,
            "required_for": _s(ws[f"B{r}"].value),
            "clause": _s(ws[f"C{r}"].value),
            "description": _s(ws[f"D{r}"].value),
            "answer": _s(ws[f"E{r}"].value),
            "remark": _s(ws[f"F{r}"].value),
            "category": _s(ws[f"G{r}"].value).upper(),
            "consumed_by": _consumed_of(ws[f"H{r}"].value),
        })
    wb.close()
    return tuple(out)


@lru_cache(maxsize=1)
def material_classes(path=None):
    """Sheet2 -- the material-classification dropdown."""
    p = Path(path) if path else XLSX
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb[CLASS_SHEET]
    out = [_s(c.value) for (c,) in ws.iter_rows(min_col=1, max_col=1) if _s(c.value)]
    wb.close()
    return tuple(out)


def block(name):
    return [r for r in rows() if r["block"] == name]


def find(block_name, sl):
    return next((r for r in rows()
                 if r["block"] == block_name and r["sl"] == str(sl)), None)


def consumed_by(key):
    """Rows feeding one downstream document -- 'tender+tec_report' or 'tender+comm_eval'."""
    return [r for r in rows() if r["consumed_by"] == key]


def by_category(cat):
    return [r for r in rows() if r["category"] == cat.upper()]


def answers():
    """The sheet's own filled-in column E, keyed "<block>:<sl>".

    The June-2026 revision ships a worked example (Goods, Capital Equipment, Technical
    Estimate, Yes-CPA is Level I, Two Bid, Line Wise, GEM-Custom bid), so the chain can
    be built with nothing hand-authored.
    """
    return {f"{r['block']}:{r['sl']}": r["answer"] for r in rows()}


def answer_of(ans, block_name, sl):
    return ans.get(f"{block_name}:{sl}", "")


def dop_level(ans=None):
    """The DOP level the indentor recorded against provisioning sl 11.

    The sample answer is "Yes-CPA is Level I". Returns e.g. "Level I", or None -- the
    value bands that would let this be *computed* are not in sampleData, so it stays a
    human-supplied fact, exactly as rules.dop_cfa_level() already treats it.
    """
    a = answer_of(ans if ans is not None else answers(), PROVISIONING, "11")
    m = re.search(r"LEVEL\s+([IVX]+)", a.upper())
    return f"Level {m.group(1)}" if m else None


def injected(ans=None):
    """Which extra authorities this case's answers pull into the approval chain.

    Each result carries the row that caused it and the sheet's own wording, so the CLI
    can show *why* somebody was added rather than just that they were.
    """
    ans = answers() if ans is None else ans
    out = []
    for spec in INJECTIONS:
        row = find(spec["block"], spec["sl"])
        if row is None:
            continue
        answer = answer_of(ans, spec["block"], spec["sl"])
        yes = is_affirmative(answer)
        fires = yes if spec["trigger"] == "affirmative" else not yes
        if not fires:
            continue
        out.append({
            "kind": spec["kind"],
            "authority": spec["authority"],
            "external": spec.get("external", False),
            "row": row["row"],
            "block": spec["block"],
            "sl": spec["sl"],
            "clause": row["clause"],
            "answer": answer,
            "trigger": spec["trigger"],
            "why": row["description"] or row["clause"],
        })
    return out


def evidence_text(row):
    """Clause + description + remark, which is where an injection's evidence must live."""
    return " | ".join(x for x in (row["clause"], row["description"], row["remark"]) if x)


def summary():
    rs = rows()
    t, c = len(by_category("T")), len(by_category("C"))
    return "\n".join([
        f"indentor checklist: {len(rs)} rows (sheet rows {FIRST_ROW}-{LAST_ROW}) from {XLSX.name}",
        f"  blocks       : {len(block(PROVISIONING))} provisioning + {len(block(TENDER))} tender",
        f"  category     : {t} technical (T), {c} commercial (C), {len(rs) - t - c} unmarked",
        f"  feeds        : {len(consumed_by('tender+tec_report'))} rows -> TEC Report, "
        f"{len(consumed_by('tender+comm_eval'))} rows -> Commercial Evaluation, "
        f"{len(consumed_by('indentor'))} indentor-only",
        f"  material cls : {len(material_classes())} options (Sheet2)",
        f"  DOP level    : {dop_level() or 'not stated'}",
    ])


if __name__ == "__main__":
    print(summary())

    print("\nauthorities this case's answers pull in")
    inj = injected()
    if not inj:
        print("  none")
    for i in inj:
        ext = "  [EXTERNAL to HAL]" if i["external"] else ""
        print(f"  sl {i['sl']:<3} {i['authority']}{ext}")
        print(f"         answer  : {i['answer'] or '(blank)'}   (trigger: {i['trigger']})")
        print(f"         because : {i['why'][:100]}")

    print("\ninjection rows NOT firing on this answer set")
    fired = {(i['block'], i['sl']) for i in inj}
    for spec in INJECTIONS:
        if (spec["block"], spec["sl"]) in fired:
            continue
        row = find(spec["block"], spec["sl"])
        a = answer_of(answers(), spec["block"], spec["sl"]) or "(blank)"
        print(f"  sl {spec['sl']:<3} {spec['authority'][:52]:<54} answer={a}")
    sys.exit(0)
