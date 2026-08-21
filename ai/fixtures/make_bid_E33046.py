"""Build a FILLED technical-bid compliance sheet for tender E-33046.

    conda run -n hal python ai/fixtures/make_bid_E33046.py

Why this exists
---------------
`sampleData/TechnicalBid E-33046.pdf` is HAL's *outgoing* form: the item row, a
12-line technical specification, 18 terms & conditions, and a vendor-details block --
all with the COMPLIANCE YES/NO and REMARKS columns blank, and no bidder named. It is
what HAL issues; nothing in sampleData shows what comes back.

That missing return leg is where the EMD decision and the whole TEC evaluation come
from, so the pipeline has never had a real input for either. This script fabricates it:
six bidders returning the same sheet, filled.

**Everything invented here is marked.** The tender reference, the item, the quantity,
the 12 specification lines and the 18 terms & conditions are quoted verbatim from the
client's PDF. The bidders, their compliance answers, their vendor details and every
price are fabricated -- vendor names are DV1..DV6 ("Dummy Vendor") so no real supplier
is implied, and each sheet carries a FABRICATED banner.

Outputs (both gitignored-adjacent, under ai/fixtures/):
  TechnicalBid_E-33046_FILLED.xlsx   one sheet per bidder + a PRICE BIDS sheet
  -> then run ai/bid_sheet.py to turn it into pipeline input.
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = Path(__file__).resolve().parent
OUT = HERE / "TechnicalBid_E-33046_FILLED.xlsx"

BANNER = ("FABRICATED TEST DATA -- bidders, compliance answers, vendor details and "
          "prices are invented. Tender ref, item, quantity, specifications and terms "
          "are quoted from sampleData/TechnicalBid E-33046.pdf.")

# -- verbatim from the client's PDF ------------------------------------------
TENDER_REF = "NK-M-1989(2544-1C)-MP-E-33046"
ENQUIRY_NO = "E-33046"
ITEM = ("250W High Bay LED Light Fitting LED High Bay Lighting for Ceiling "
        "Suspension Type, 250W, Specification Details as per enclosure")
QTY, UOM = 2150, "Nos."

SPECS = [
    (1, "Luminaire System Wattage (in Watts)", "250 WATTS"),
    (2, "LED Power rating corresponding minimum lumen output", "130 lm/W"),
    (3, "Color temperature", "5000K - 6500K"),
    (4, "Operating Input Voltage range (in Volts)", "140V - 270V,  50Hz"),
    (5, "Housing", "Die-cast aluminium with heat sink"),
    (6, "Shape of luminaire", "Round"),
    (7, "Ingress Protection (IP) Rating of Luminaire", "IP 66"),
    (8, "Operating Temperature range (in deg. C)", "-10°C to 45°C"),
    (9, "Life Span", "> 50000h"),
    (10, "Standards Compliance",
     "BIS-107294 (CRS NUMBER); IS 15885 (Control Gear); IS 16105; IS 16106; "
     "IS 10322 (LED luminaire)"),
    (11, "Mounting", "Hook, Pendant, or adjustable"),
    (12, "Rated Supply", "240V AC, 50 Hz"),
]

TERMS = [
    (1, "Final Inspection",
     "Final Inspection will be done at HAL site after supply, installation and "
     "successful Functional test (Installation & Functional Testing to be done by HAL)."),
    (2, "Delivery Schedule",
     "To be supplied by the Vendor within 60 Days from the date of placement of "
     "Purchase Order from HAL"),
    (3, "Warranty", "Minimum 1 year warranty from the date of final acceptance at HAL's site."),
    (4, "Tender Submission Due Date", "As per GeM portal"),
    (5, "Tender Evaluation", "Line wise"),
    (6, "Submission Of Tender In Bid System.", "Two Bid"),
    (7, "GST Registration", "Vendor should have GST registration number."),
    (8, "Conditions of the contract", "Applicable and enclosed as per Technical Specification"),
    (9, "Payment term",
     "No advance payment will be made. Our standard payment term is “Full payment "
     "within 30 days after receipt and acceptance of material at HAL premises”."),
    (10, "Liquidated damages",
     "Delivery quoted should be firm. In case of delay in delivery of material or delay "
     "in commissioning, HAL reserves the right to collect the sum of 0.5% per week or "
     "part thereof, subject to a maximum of 10% as HAL’s claim towards liquidated "
     "damages on undelivered / unexecuted part of the order for SITC of Equipment."),
    (11, "Taxes & Excise Duties",
     "Please intimate us whether any taxes or duties are leviable for the items offered. "
     "If so, percentage thereof may please be specified in the quotation. HAL will assume "
     "that the rates quoted are inclusive of GST. CESS Not Applicable and shall not be "
     "reimbursed by HAL."),
    (12, "Incomplete Offers",
     "HAL reserves the Right to reject, either fully or partially, any offer without "
     "assigning reasons."),
    (13, "Bank Details",
     "The Bidder is requested to upload the Original cancelled cheque copy and Original "
     "Bank Mandate format. Bank Account should be same in Tender & GeM account."),
    (14, "Compliance Certificate",
     "The Supplier should furnish a compliance certificate as per the format enclosed "
     "indicating compliance to the tender specifications. Wherever offer deviates from "
     "the tender specifications, same should be brought out clearly indicating the "
     "nature / extent of deviation(s)."),
    (15, "Declaration from Bidder on being Debarred, Suspended, Disqualified, Removed, "
         "Business Dealings Banned etc. by, HAL, GoI & its associated organizations",
     "Declaration to be signed Stamped & submitted by Bidder as per enclosed "
     "Annexure - I along with Bid Documents"),
    (16, "Make In India (MIL) Clause",
     "LOCAL CONTENT AS PER PUBLIC PROCURMENT (PREFERENCE TO MAKE IN INDIA) ORDER- 2017 "
     "( MLC %), Class-I Local Supplier: Equal to or more than 50% & Class-II Local "
     "Supplier: More than 20% but less than 50%."),
    (17, "Land Boundry Sharing Declaration",
     "The Bidder has to submit the applicable Land Boundry Sharing Declaration "
     "certificate as per the attached formats. (Either Type I or Type II)."),
    (18, "Special Note 1",
     "Bidder has submitted below mentioned documents: 1) Original Cancelled Cheque "
     "2) Original Bank Mandate form certified by Bank 3) GST Certificate"),
]

VENDOR_FIELDS = [
    "Name of Firm/Organization", "Registered Address", "GST No", "PAN No",
    "Phone No/Mobile No", "e-Mail Address", "Bank Account No (To be same GeM & invoice)",
    "IFSC Code of Bank Account & Bank Name (To be same GeM & invoice)",
    "Bank Branch Address", "HAL Vendor Code (if any)",
    "Nature of Firm/Organization (Manufacturer/OEM/Stockist/Retailer/Authorised Dealer, etc.)",
    "MSME (Yes or NO), Enclose relevant certificate",
    "Classification of Manufacturer (Class I / II / III), Enclose relevant certificate",
    "Start Up (Yes or NO), Enclose relevant certificate",
    "SC/ST/Women Entrepreneur, (Specify)",
    "Boarded on TReDS Platform : Yes or No, (Specify)",
]

# -- FABRICATED from here down ----------------------------------------------
# `spec_no` lists the specification sl nos the bidder does NOT meet, with the value
# they actually offered -- this is what a TEC report cites when it rejects an offer.
# `emd`: "Paid" or a waiver claim. A waiver is only valid if the bidder manufactures
# the offered product in the relevant NIC category (rules.emd_waiver).
BIDDERS = [
    {
        "id": "DV1", "name": "Dummy Vendor 1 Lighting Pvt. Ltd.",
        "address": "Plot 14, MIDC Ambad, Nashik, Maharashtra 422010",
        "gst": "27AAACD1111A1Z1", "pan": "AAACD1111A",
        "phone": "+91-9000000001", "email": "tenders@dv1.example",
        "bank_ac": "500100100001", "ifsc": "DUMB0000001 / Dummy Bank of India",
        "branch": "Ambad Industrial Area, Nashik", "hal_code": "V-DV1-0001",
        "nature": "Manufacturer", "msme": "Yes - Small", "mfr_class": "Class I",
        "startup": "No", "social": "None", "treds": "Yes",
        "udyam": "UDYAM-MH-19-0100001", "nic": "27400",
        "emd": "Waiver sought - MSE (Udyam enclosed), manufacturer of LED luminaires",
        "mlc_pct": 62, "mlc_class": "Class-I",
        "spec_no": [],
        "unit_basic": 4950,
    },
    {
        "id": "DV2", "name": "Dummy Vendor 2 Illumination Systems",
        "address": "Sector 8, Industrial Estate, Pune, Maharashtra 411019",
        "gst": "27AAACD2222B1Z2", "pan": "AAACD2222B",
        "phone": "+91-9000000002", "email": "sales@dv2.example",
        "bank_ac": "500100100002", "ifsc": "DUMB0000002 / Dummy Bank of India",
        "branch": "Bhosari, Pune", "hal_code": "",
        "nature": "Manufacturer", "msme": "Yes - Micro", "mfr_class": "Class I",
        "startup": "No", "social": "Women Entrepreneur", "treds": "No",
        "udyam": "UDYAM-MH-26-0100002", "nic": "27400",
        "emd": "Waiver sought - MSE (Udyam enclosed), manufacturer of LED luminaires",
        "mlc_pct": 55, "mlc_class": "Class-I",
        "spec_no": [(2, "110 lm/W offered against 130 lm/W required"),
                    (7, "IP 54 offered against IP 66 required")],
        "unit_basic": 4310,
    },
    {
        "id": "DV3", "name": "Dummy Vendor 3 Electricals Ltd.",
        "address": "GIDC Estate, Vadodara, Gujarat 390010",
        "gst": "24AAACD3333C1Z3", "pan": "AAACD3333C",
        "phone": "+91-9000000003", "email": "bids@dv3.example",
        "bank_ac": "500100100003", "ifsc": "DUMB0000003 / Dummy Bank of India",
        "branch": "Makarpura, Vadodara", "hal_code": "V-DV3-0007",
        "nature": "Manufacturer / OEM", "msme": "No - Non-MSME", "mfr_class": "Class I",
        "startup": "No", "social": "None", "treds": "Yes",
        "udyam": "N/A - Non-MSME", "nic": "27400",
        "emd": "Paid - SB Collect ref DUMMYSBI0003 dt 12.06.2026",
        "mlc_pct": 71, "mlc_class": "Class-I",
        "spec_no": [],
        "unit_basic": 5400,
    },
    {
        "id": "DV4", "name": "Dummy Vendor 4 Lumen Technologies",
        "address": "Peenya Industrial Area, Bengaluru, Karnataka 560058",
        "gst": "29AAACD4444D1Z4", "pan": "AAACD4444D",
        "phone": "+91-9000000004", "email": "tender@dv4.example",
        "bank_ac": "500100100004", "ifsc": "DUMB0000004 / Dummy Bank of India",
        "branch": "Peenya, Bengaluru", "hal_code": "",
        "nature": "Manufacturer", "msme": "Yes - Small", "mfr_class": "Class II",
        "startup": "Yes - DPIIT recognised", "social": "None", "treds": "No",
        "udyam": "UDYAM-KR-03-0100004", "nic": "27400",
        "emd": "Waiver sought - MSE + Start-up (DPIIT certificate enclosed)",
        "mlc_pct": 34, "mlc_class": "Class-II",
        "spec_no": [(9, "30000h offered against > 50000h required"),
                    (10, "BIS CRS registration not enclosed")],
        "unit_basic": 4180,
    },
    {
        "id": "DV5", "name": "Dummy Vendor 5 Trading Company",
        "address": "Chandni Chowk, New Delhi 110006",
        "gst": "07AAACD5555E1Z5", "pan": "AAACD5555E",
        "phone": "+91-9000000005", "email": "info@dv5.example",
        "bank_ac": "500100100005", "ifsc": "DUMB0000005 / Dummy Bank of India",
        "branch": "Chandni Chowk, New Delhi", "hal_code": "",
        "nature": "Stockist / Retailer", "msme": "Yes - Micro", "mfr_class": "Class II",
        "startup": "No", "social": "None", "treds": "No",
        # NIC 46592 is wholesale trading, not manufacture of lighting equipment (27400).
        # The waiver claim therefore fails at the EMD stage -- the same test the NVB
        # case applied to Anika Steel and Hazzale Blue Trading.
        "udyam": "UDYAM-DL-06-0100005", "nic": "46592",
        "emd": "Waiver sought - MSE (Udyam enclosed)",
        "mlc_pct": 22, "mlc_class": "Class-II",
        "spec_no": [(2, "120 lm/W offered against 130 lm/W required")],
        "unit_basic": 4020,
    },
    {
        "id": "DV6", "name": "Dummy Vendor 6 Power & Light Corp.",
        "address": "Sahibabad Industrial Area, Ghaziabad, Uttar Pradesh 201010",
        "gst": "09AAACD6666F1Z6", "pan": "AAACD6666F",
        "phone": "+91-9000000006", "email": "contact@dv6.example",
        "bank_ac": "500100100006", "ifsc": "DUMB0000006 / Dummy Bank of India",
        "branch": "Sahibabad, Ghaziabad", "hal_code": "",
        "nature": "Authorised Dealer", "msme": "No - Non-MSME", "mfr_class": "Class II",
        "startup": "No", "social": "None", "treds": "No",
        "udyam": "N/A - Non-MSME", "nic": "46592",
        # No EMD and no valid ground for waiver -> rejected at the EMD stage.
        "emd": "Not paid; waiver claimed without supporting document",
        "mlc_pct": 18, "mlc_class": "Not eligible (below 20%)",
        "spec_no": [],
        "unit_basic": 3990,
    },
]

# The case around the bids -- all fabricated.
GST_PCT = 0.18
ESTIMATE_UNIT_BASIC = 4200            # provisioning estimate, per unit, excl. GST
LPP_UNIT_BASIC = 4100                 # last purchase price, per unit, excl. GST
LPP_CONTRACT = "HAL/NK/IMM/LED/2024/0451 dt 22.08.2024"
COUNTER_UNIT_BASIC = 4400             # what L1 agreed to after negotiation
RA_STATUS = "Reverse Auction floated on GeM; no bidder participated"
CAR_NO, CAR_DATE = "CAR/26/118", "11 May 2026"
TENDER_NO, TENDER_DATE = "GEM/2026/B/7729104", "05-06-2026"
DUE_DATE = "26-06-2026"

# -- styling ----------------------------------------------------------------
THIN = Side(style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR = PatternFill("solid", fgColor="DDE7F0")
WARN = PatternFill("solid", fgColor="FFF2CC")
TITLE = Font(bold=True, size=12)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _row(ws, r, values, fill=None, bold=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BOX
        c.alignment = WRAP
        if fill:
            c.fill = fill
        if bold:
            c.font = BOLD
    return r + 1


def _compliance(bidder, sl):
    """The bidder's YES/NO and remark against one specification line."""
    for no, why in bidder["spec_no"]:
        if no == sl:
            return "NO", why
    return "YES", "Complied"


def build_bidder_sheet(wb, b):
    ws = wb.create_sheet(b["id"])
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 46

    r = 1
    ws.cell(row=r, column=1, value="HINDUSTAN AERONAUTICS LIMITED - Aircraft Overhaul "
                                   "Division, NASIK, OJHAR(MIG), NASIK-422 207, INDIA.").font = TITLE
    r += 1
    ws.cell(row=r, column=1, value="TECHNICAL BID COMPLIANCE SHEET").font = TITLE
    r += 1
    ws.cell(row=r, column=1, value=f"Tender Ref: {TENDER_REF}").font = BOLD
    r += 1
    ws.cell(row=r, column=1, value=f"NAME OF THE BIDDER: {b['name']}").font = BOLD
    r += 1
    c = ws.cell(row=r, column=1, value=BANNER)
    c.fill, c.alignment = WARN, WRAP
    r += 2

    r = _row(ws, r, ["Sr .No", "DESCRIPTION", "QTY", "UNIT OF MEASUREMENT",
                     "COMPLIANCE YES/NO", "REMARKS IF ANY"], fill=HDR, bold=True)
    r = _row(ws, r, [1, ITEM, QTY, UOM, "YES",
                     f"Offered as per enclosure. MLC {b['mlc_pct']}% ({b['mlc_class']})."])
    r += 1

    ws.cell(row=r, column=1, value="TECHNICAL SPECIFICATIONS").font = BOLD
    r += 1
    r = _row(ws, r, ["Sr .No", "Specification", "HAL Requirement", "",
                     "COMPLIANCE YES/NO", "REMARKS IF ANY"], fill=HDR, bold=True)
    r = _row(ws, r, ["A)", "TECHNICAL SPECIFICATION FOR LED HIGHBAY FITTINGS",
                     "", "", "", ""], bold=True)
    for sl, name, req in SPECS:
        yn, remark = _compliance(b, sl)
        r = _row(ws, r, [sl, name, req, "", yn, remark])
    r += 1

    ws.cell(row=r, column=1, value="OTHER APPLICABLE TERMS AND CONDITIONS").font = BOLD
    r += 1
    r = _row(ws, r, ["Sl. NO.", "TERMS & CONDITION", "DESCRIPTION",
                     "COMPLIANCE YES/NO", "REMARKS IF ANY", ""], fill=HDR, bold=True)
    for sl, term, desc in TERMS:
        remark = "Accepted"
        if sl == 16:
            remark = f"MLC {b['mlc_pct']}% - {b['mlc_class']}. Declaration enclosed."
        elif sl == 17:
            remark = "Type-I declaration enclosed (not a land-border country)."
        elif sl == 3:
            remark = "12 months from final acceptance."
        elif sl == 2:
            remark = "60 days from PO accepted."
        r = _row(ws, r, [sl, term, desc, "YES", remark, ""])
    r += 1

    ws.cell(row=r, column=1, value="VENDOR DETAILS (filled by Bidder)").font = BOLD
    r += 1
    values = [b["name"], b["address"], b["gst"], b["pan"], b["phone"], b["email"],
              b["bank_ac"], b["ifsc"], b["branch"], b["hal_code"], b["nature"],
              b["msme"], b["mfr_class"], b["startup"], b["social"], b["treds"]]
    for i, (label, val) in enumerate(zip(VENDOR_FIELDS, values), 1):
        r = _row(ws, r, [i, label, val, "", "", ""])
    r += 1

    ws.cell(row=r, column=1, value="EMD / UDYAM (filled by Bidder)").font = BOLD
    r += 1
    r = _row(ws, r, ["", "Udyam Registration No.", b["udyam"], "", "", ""])
    r = _row(ws, r, ["", "NIC code of the offered product", b["nic"], "", "", ""])
    r = _row(ws, r, ["", "EMD status / waiver claimed", b["emd"], "", "", ""])
    r += 1
    ws.cell(row=r, column=5, value="BIDDER SIGN & STAMP").font = BOLD
    return ws


def build_price_sheet(wb):
    """The price bid -- a separate envelope, opened only after the TEC stage."""
    ws = wb.create_sheet("PRICE BIDS")
    for col, w in zip("ABCDEFG", (10, 40, 12, 16, 18, 18, 20)):
        ws.column_dimensions[col].width = w
    r = 1
    ws.cell(row=r, column=1, value=f"PRICE BIDS - Tender {TENDER_REF}").font = TITLE
    r += 1
    c = ws.cell(row=r, column=1, value=BANNER)
    c.fill, c.alignment = WARN, WRAP
    r += 2
    ws.cell(row=r, column=1, value="Two-bid tender: this envelope is opened only after "
                                   "the TEC has cleared the technical bids.").font = BOLD
    r += 2
    r = _row(ws, r, ["Bidder", "Name", "Qty", "Unit rate (INR, basic)",
                     "Basic value (INR)", f"GST @ {int(GST_PCT * 100)}%",
                     "Landed value (INR)"], fill=HDR, bold=True)
    for b in BIDDERS:
        basic = b["unit_basic"] * QTY
        gst = round(basic * GST_PCT, 2)
        r = _row(ws, r, [b["id"], b["name"], QTY, b["unit_basic"],
                         basic, gst, round(basic + gst, 2)])
    r += 1
    est_basic = ESTIMATE_UNIT_BASIC * QTY
    lpp_basic = LPP_UNIT_BASIC * QTY
    for label, unit, basic in [
        ("Provisioning estimate", ESTIMATE_UNIT_BASIC, est_basic),
        ("Last purchase price (LPP)", LPP_UNIT_BASIC, lpp_basic),
        ("Counter-offer accepted after PNC", COUNTER_UNIT_BASIC, COUNTER_UNIT_BASIC * QTY),
    ]:
        gst = round(basic * GST_PCT, 2)
        r = _row(ws, r, [label, "", QTY, unit, basic, gst, round(basic + gst, 2)],
                 bold=True)
    r += 1
    r = _row(ws, r, ["Reverse Auction", RA_STATUS, "", "", "", "", ""])
    r = _row(ws, r, ["LPP contract", LPP_CONTRACT, "", "", "", "", ""])
    return ws


def build_readme(wb):
    ws = wb.create_sheet("READ ME", 0)
    ws.column_dimensions["A"].width = 118
    lines = [
        ("FILLED technical-bid compliance sheets - tender " + TENDER_REF, TITLE),
        ("", None),
        (BANNER, None),
        ("", None),
        ("Quoted verbatim from sampleData/TechnicalBid E-33046.pdf:", BOLD),
        ("   tender reference, item description, quantity (2150 Nos.),", None),
        ("   the 12 technical specification lines and their required values,", None),
        ("   the 18 other applicable terms and conditions and their descriptions,", None),
        ("   the 16 vendor-detail field labels.", None),
        ("", None),
        ("Fabricated for testing:", BOLD),
        ("   the six bidders DV1-DV6 and every value they entered,", None),
        ("   all COMPLIANCE YES/NO answers and REMARKS,", None),
        ("   Udyam numbers, NIC codes, EMD status, MLC percentages,", None),
        ("   every price, the provisioning estimate, the LPP and the counter-offer.", None),
        ("", None),
        ("No real supplier is named or implied. Names are DV1-DV6, addresses are", None),
        ("generic industrial estates, e-mail domains are .example (RFC 2606), and", None),
        ("bank/GST/PAN identifiers are patterned placeholders.", None),
        ("", None),
        ("Regenerate:  conda run -n hal python ai/fixtures/make_bid_E33046.py", BOLD),
        ("Load it:     conda run -n hal python ai/bid_sheet.py", BOLD),
    ]
    for i, (text, font) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if font:
            c.font = font
    ws["A3"].fill = WARN
    return ws


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_readme(wb)
    for b in BIDDERS:
        build_bidder_sheet(wb, b)
    build_price_sheet(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    print(f"wrote {OUT.relative_to(HERE.parents[1])}")
    print(f"  {len(BIDDERS)} bidder sheets + PRICE BIDS + READ ME")
    print(f"  tender {TENDER_REF}, item 250W High Bay LED, qty {QTY} {UOM}")
    print(f"  spec lines quoted: {len(SPECS)}   terms quoted: {len(TERMS)}")
    noncompliant = [b["id"] for b in BIDDERS if b["spec_no"]]
    print(f"  bidders with spec deviations: {', '.join(noncompliant)}")
    print("\n  FABRICATED: bidders, compliance answers, vendor details, all prices.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
